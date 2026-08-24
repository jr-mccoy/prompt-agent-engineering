#!/usr/bin/env python3
"""
build_workbook.py - Assemble extracted (and optionally categorized/flagged)
transaction CSVs into a polished Excel workbook for an attorney or reviewer.

Reads one or more <stem>_transactions.csv files (and, when present, their
<stem>_reconciliation.json and <stem>_meta.json siblings) and produces a single
.xlsx workbook with these tabs:
    * Transactions   - every row from every statement, with a source column
    * Summary        - per-statement counts/credits/debits/net + by-category totals
    * Reconciliation - per-statement verification result (balance + coverage)
    * Flags          - only the rows that carry a divorce/custody flag (if any)

Requires openpyxl for .xlsx output. If openpyxl is not installed, the script
writes a combined CSV instead and tells you how to install it (graceful fallback).

Usage:
    build_workbook.py [CSV ...] [--inputs DIR] --out FILE.xlsx

Options:
    --inputs DIR   Directory to scan for *_transactions.csv (in addition to CSV args)
    --out FILE     Output workbook path (default: output/master_workbook.xlsx)
    --title TEXT   Title used on the Summary tab (default: "Financial Records")

Examples:
    build_workbook.py output/chase_jan_transactions.csv --out output/chase_jan.xlsx
    build_workbook.py --inputs output/ --out output/master_workbook.xlsx

Exit codes:
    0 success (workbook or fallback CSV written)
    2 no input CSVs found
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

SCHEMA = [
    "source_file", "statement_period", "account_id", "date",
    "description_raw", "description_clean", "amount", "direction",
    "running_balance", "category", "subcategory", "flags",
    "confidence", "needs_review", "notes",
]


def _to_float(s):
    try:
        return float((s or "").strip())
    except (ValueError, AttributeError):
        return 0.0


def collect_csvs(args_csvs, inputs_dir) -> list[Path]:
    paths: list[Path] = [Path(p) for p in args_csvs]
    if inputs_dir:
        paths.extend(sorted(Path(inputs_dir).glob("*_transactions.csv")))
    # De-dup while preserving order.
    seen, out = set(), []
    for p in paths:
        if p.exists() and str(p) not in seen:
            seen.add(str(p))
            out.append(p)
    return out


def load(csv_path: Path):
    with csv_path.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    stem = csv_path.stem.replace("_transactions", "")
    recon_path = csv_path.with_name(f"{stem}_reconciliation.json")
    recon = json.loads(recon_path.read_text(encoding="utf-8")) if recon_path.exists() else {}
    return rows, recon


def build_summary(all_rows):
    by_src: dict[str, dict] = {}
    by_cat: dict[str, float] = {}
    for r in all_rows:
        src = r.get("source_file", "?")
        amt = _to_float(r.get("amount"))
        s = by_src.setdefault(src, {"count": 0, "credits": 0.0, "debits": 0.0,
                                     "period": r.get("statement_period", "")})
        s["count"] += 1
        if amt >= 0:
            s["credits"] += amt
        else:
            s["debits"] += amt
        cat = r.get("category") or "UNCATEGORIZED"
        by_cat[cat] = by_cat.get(cat, 0.0) + amt
    return by_src, by_cat


def write_csv_fallback(all_rows, out_path: Path):
    fallback = out_path.with_suffix(".csv")
    with fallback.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=SCHEMA)
        writer.writeheader()
        for r in all_rows:
            writer.writerow({k: r.get(k, "") for k in SCHEMA})
    print(f"[workbook] openpyxl not installed -> wrote combined CSV: {fallback}")
    print("[workbook] For the formatted .xlsx workbook: pip install openpyxl")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build an Excel workbook from transaction CSVs.")
    parser.add_argument("csvs", nargs="*", help="transactions CSV files")
    parser.add_argument("--inputs", default="", help="directory to scan for *_transactions.csv")
    parser.add_argument("--out", default="output/master_workbook.xlsx")
    parser.add_argument("--title", default="Financial Records")
    args = parser.parse_args()

    csv_paths = collect_csvs(args.csvs, args.inputs)
    if not csv_paths:
        print("[workbook] ERROR: no input transactions CSVs found.")
        return 2

    all_rows, recons = [], []
    for p in csv_paths:
        rows, recon = load(p)
        all_rows.extend(rows)
        recons.append((p.name, recon))

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        write_csv_fallback(all_rows, out_path)
        return 0

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    flag_fill = PatternFill("solid", fgColor="FCE4D6")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    # --- Transactions tab ---
    ws = wb.active
    ws.title = "Transactions"
    ws.append(SCHEMA)
    for r in all_rows:
        row = [r.get(k, "") for k in SCHEMA]
        ws.append(row)
        if (r.get("flags") or "").strip():
            for c in range(1, len(SCHEMA) + 1):
                ws.cell(row=ws.max_row, column=c).fill = flag_fill
    style_header(ws, len(SCHEMA))
    for i, _ in enumerate(SCHEMA, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # --- Summary tab ---
    by_src, by_cat = build_summary(all_rows)
    ws2 = wb.create_sheet("Summary")
    ws2.append([args.title])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["Statement", "Period", "Transactions", "Credits (in)", "Debits (out)", "Net"])
    hdr_row = ws2.max_row
    for src, s in by_src.items():
        ws2.append([src, s["period"], s["count"], round(s["credits"], 2),
                    round(s["debits"], 2), round(s["credits"] + s["debits"], 2)])
    for c in range(1, 7):
        ws2.cell(row=hdr_row, column=c).fill = header_fill
        ws2.cell(row=hdr_row, column=c).font = header_font
    ws2.append([])
    ws2.append(["By category", "Net amount"])
    cat_hdr = ws2.max_row
    for cat, amt in sorted(by_cat.items(), key=lambda kv: kv[1]):
        ws2.append([cat, round(amt, 2)])
    for c in range(1, 3):
        ws2.cell(row=cat_hdr, column=c).fill = header_fill
        ws2.cell(row=cat_hdr, column=c).font = header_font
    for col in ("A", "B", "C", "D", "E", "F"):
        ws2.column_dimensions[col].width = 20

    # --- Reconciliation tab ---
    ws3 = wb.create_sheet("Reconciliation")
    ws3.append(["Statement", "Result", "Balance", "Coverage", "Errors", "# Warnings"])
    for c in range(1, 7):
        ws3.cell(row=1, column=c).fill = header_fill
        ws3.cell(row=1, column=c).font = header_font
    for name, recon in recons:
        if not recon:
            ws3.append([name, "NOT VERIFIED", "-", "-", "no reconciliation report", 0])
            continue
        ws3.append([
            name,
            recon.get("result", "?"),
            recon.get("layer1_balance", {}).get("status", "?"),
            recon.get("coverage_check", {}).get("status", "?"),
            "; ".join(recon.get("errors", [])) or "-",
            len(recon.get("warnings", [])),
        ])
    ws3.freeze_panes = "A2"
    for col in ("A", "B", "C", "D", "E", "F"):
        ws3.column_dimensions[col].width = 22

    # --- Flags tab (only if any row is flagged) ---
    flagged = [r for r in all_rows if (r.get("flags") or "").strip()]
    if flagged:
        ws4 = wb.create_sheet("Flags")
        cols = ["source_file", "date", "description_clean", "amount", "category", "flags", "needs_review", "notes"]
        ws4.append(cols)
        for c in range(1, len(cols) + 1):
            ws4.cell(row=1, column=c).fill = header_fill
            ws4.cell(row=1, column=c).font = header_font
        for r in flagged:
            ws4.append([r.get(k, "") for k in cols])
        ws4.freeze_panes = "A2"
        for i in range(1, len(cols) + 1):
            ws4.column_dimensions[get_column_letter(i)].width = 22

    wb.save(out_path)
    print(f"[workbook] OK: {len(all_rows)} transactions from {len(csv_paths)} statement(s) -> {out_path}")
    print(f"[workbook] tabs: Transactions, Summary, Reconciliation"
          + (", Flags" if flagged else ""))
    return 0


if __name__ == "__main__":
    sys.exit(main())
